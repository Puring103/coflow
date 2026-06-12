"""
Convert Luban-format xlsx to a coflow-compatible xlsx.

Luban convention:
  Row 1: control col + ##var + field names
  Row 2: control col + ##type + types
  Row 3: control col + ##   + chinese descriptions
  Row 4+: data, control col empty for normal rows; "##" comments-out a row

Coflow convention:
  Row 1 = field names; row 2+ = data; arrays separated by `|`.

Coflow already accepts:
  - enum-typed `@id` (so we keep the enum NAME as-is, no int conversion)
  - `0`/`1` for bool fields (the cell parser is permissive)
  - `@expand` parent fields whose adjacent columns are independent cells

The only structural rewrite the converter still does:
  - Drop the Luban control column and the two extra header rows (##type, ##).
  - Drop placeholder rows (only id filled, all other cells empty).
  - Rewrite array `,` separators to `|`.
  - Fill in inner field names underneath an `@expand` parent header
    (Luban merges those header cells; coflow needs them visible).
"""
from __future__ import annotations

import os
import re
import sys
import zipfile
from openpyxl import Workbook

sys.stdout.reconfigure(encoding="utf-8")


SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_DIR = os.path.dirname(SCRIPT_DIR)
SRC_XLSX = os.path.normpath(
    os.path.join(
        PROJECT_DIR,
        "..",
        "..",
        "..",
        "HumanPark",
        "Assets",
        "Plugins",
        "Luban",
        "Datas",
        "Configs.xlsx",
    )
)
DST_XLSX = os.path.join(PROJECT_DIR, "data", "Configs.coflow.xlsx")


SHEETS = [
    ("生物残骸", 2),
    ("地块tag", 3),
    ("地块", 4),
    ("物质", 5),
    ("属性", 6),
    ("能力", 7),
    ("基因", 8),
    ("表皮", 10),
    ("阶段", 11),
    ("功能", 12),
]


# Sheet -> array column headers whose `,` separators (Luban convention)
# need to be rewritten to `|` (coflow array separator).
ARRAY_COLUMNS = {
    "阶段": {"unlockFeatureList", "unlockGeneIdList"},
    "生物残骸": {"generateTerrainTypes"},
}




C_TAG = re.compile(r"<c\s+([^>/]*?)(?:/>|>(.*?)</c>)", flags=re.S)
ATTR = re.compile(r'(\w+)="([^"]*)"')
V_TAG = re.compile(r"<v>(.*?)</v>", flags=re.S)
T_TAG = re.compile(r"<t[^>]*>(.*?)</t>", flags=re.S)
IS_TAG = re.compile(r"<is>(.*?)</is>", flags=re.S)


def col_letter(ref: str) -> str:
    return re.match(r"([A-Z]+)", ref).group(1)


def col_index(letter: str) -> int:
    n = 0
    for c in letter:
        n = n * 26 + (ord(c) - ord("A") + 1)
    return n


def parse_row(content: str, sst: list[str]) -> dict[int, object]:
    row: dict[int, object] = {}
    for m in C_TAG.finditer(content):
        attrs = dict(ATTR.findall(m.group(1)))
        body = m.group(2) or ""
        ref = attrs.get("r", "")
        if not ref:
            continue
        letter = col_letter(ref)
        t = attrs.get("t", "n")
        if t == "s":
            v = V_TAG.search(body)
            val = sst[int(v.group(1))] if v else ""
        elif t == "inlineStr":
            inner = IS_TAG.search(body)
            val = "".join(T_TAG.findall(inner.group(1))) if inner else ""
        elif t == "b":
            v = V_TAG.search(body)
            val = bool(int(v.group(1))) if v else False
        else:
            v = V_TAG.search(body)
            val = v.group(1) if v else ""
        row[col_index(letter)] = val
    return row


def read_luban_sheet(z: zipfile.ZipFile, sst: list[str], sheet_path: str) -> dict[int, dict]:
    with z.open(sheet_path) as f:
        xml = f.read().decode("utf-8")
    body = re.search(r"<sheetData>(.*?)</sheetData>", xml, flags=re.S)
    if not body:
        return {}
    rows_xml = re.findall(r'<row[^>]*r="(\d+)"[^>]*>(.*?)</row>', body.group(1), flags=re.S)
    rows = {}
    for r_idx, content in rows_xml:
        rows[int(r_idx)] = parse_row(content, sst)
    return rows


def coerce_value(v: object) -> object:
    if v == "" or v is None:
        return None
    if isinstance(v, bool):
        return v
    if isinstance(v, str):
        try:
            if "." in v or "e" in v.lower():
                return float(v)
            return int(v)
        except ValueError:
            return v
    return v




def main() -> None:
    if not os.path.exists(SRC_XLSX):
        print(f"missing source xlsx: {SRC_XLSX}", file=sys.stderr)
        sys.exit(1)

    with zipfile.ZipFile(SRC_XLSX) as z:
        with z.open("xl/sharedStrings.xml") as f:
            ss = f.read().decode("utf-8")
        sst = [
            "".join(T_TAG.findall(s))
            for s in re.findall(r"<si>(.*?)</si>", ss, flags=re.S)
        ]

        wb = Workbook()
        wb.remove(wb.active)

        for name, idx in SHEETS:
            rows = read_luban_sheet(z, sst, f"xl/worksheets/sheet{idx}.xml")
            if not rows:
                continue
            header = rows.get(1, {})
            ws = wb.create_sheet(title=name)
            max_col = max((max(r.keys()) for r in rows.values() if r), default=0)
            # Determine the rightmost column that carries actual data on any
            # row. Trailing fully-empty columns are dropped; interior empty
            # header cells are kept so that @expand parent columns retain
            # their adjacent positional slots even though Luban left those
            # header cells blank under the merged region.
            out_cols = list(range(2, max_col + 1))
            # Drop trailing columns whose header is empty AND whose data is
            # entirely empty across the sheet — these are stray empty columns
            # past the real schema.
            while out_cols:
                tail = out_cols[-1]
                if header.get(tail) in ("", None) and all(
                    rows[r].get(tail, "") in ("", None) for r in rows
                ):
                    out_cols.pop()
                else:
                    break
            # Also drop leading-blank columns past column 1 (control col) that
            # have no header AND no data. (Almost never triggers for Luban.)
            while len(out_cols) > 1 and header.get(out_cols[0]) in ("", None) and all(
                rows[r].get(out_cols[0], "") in ("", None) for r in rows
            ):
                out_cols.pop(0)
            field_names = [str(header.get(c) or "") for c in out_cols]
            for out_idx, field in enumerate(field_names, start=1):
                ws.cell(row=1, column=out_idx, value=field)

            array_cols = ARRAY_COLUMNS.get(name, set())

            for r in sorted(rows.keys()):
                if r < 4:
                    continue
                src_row = rows[r]
                ctrl = (src_row.get(1, "") or "").strip()
                if isinstance(ctrl, str) and ctrl.startswith("##"):
                    continue
                # Skip rows that only have an `id` cell with everything else
                # blank — Luban silently tolerates these placeholder rows but
                # coflow would flag every missing required field.
                non_id_filled = any(
                    src_row.get(c, "") not in ("", None)
                    for c in out_cols[1:]
                )
                if not non_id_filled:
                    continue
                target_row = ws.max_row + 1
                for out_idx, src_col in enumerate(out_cols, start=1):
                    field = field_names[out_idx - 1]
                    raw = src_row.get(src_col, "")
                    if raw == "" or raw is None:
                        continue
                    if field in array_cols and isinstance(raw, str):
                        ws.cell(
                            row=target_row,
                            column=out_idx,
                            value=raw.replace(",", "|"),
                        )
                        continue
                    ws.cell(row=target_row, column=out_idx, value=coerce_value(raw))

        os.makedirs(os.path.dirname(DST_XLSX), exist_ok=True)
        wb.save(DST_XLSX)
        print(f"wrote {DST_XLSX}")


if __name__ == "__main__":
    main()
